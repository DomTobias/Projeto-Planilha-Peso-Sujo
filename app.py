from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import os
import pandas as pd
from io import BytesIO
from calendar import monthrange
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

app = Flask(__name__)
DATABASE = 'controle_peso.db'

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    if not os.path.exists(DATABASE):
        with app.app_context():
            db = get_db()
            with open('schema.sql', mode='r') as f:
                db.cursor().executescript(f.read())
            db.commit()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/dados/<data_alvo>')
def get_dados(data_alvo):
    db = get_db()
    db.execute('INSERT OR IGNORE INTO registros_dia (data) VALUES (?)', (data_alvo,))
    db.commit()
    
    query = '''
        SELECT s.nome as setor, h.descricao as horario, p.peso
        FROM pesos p
        JOIN setores s ON p.setor_id = s.id
        JOIN horarios h ON p.horario_id = h.id
        JOIN registros_dia r ON p.registro_id = r.id
        WHERE r.data = ?
    '''
    rows = db.execute(query, (data_alvo,)).fetchall()
    return jsonify([dict(row) for row in rows])

@app.route('/api/save', methods=['POST'])
def save():
    payload = request.json
    db = get_db()
    try:
        reg_id = db.execute('SELECT id FROM registros_dia WHERE data = ?', (payload['data'],)).fetchone()['id']
        
        db.execute('INSERT OR IGNORE INTO setores (nome) VALUES (?)', (payload['setor'],))
        set_id = db.execute('SELECT id FROM setores WHERE nome = ?', (payload['setor'],)).fetchone()['id']
        
        hor_id = db.execute('SELECT id FROM horarios WHERE descricao = ?', (payload['horario'],)).fetchone()['id']

        db.execute('''
            INSERT INTO pesos (registro_id, setor_id, horario_id, peso)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(registro_id, setor_id, horario_id) DO UPDATE SET peso = excluded.peso
        ''', (reg_id, set_id, hor_id, payload['peso']))
        
        db.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 400

@app.route('/api/consolidado/<ano>/<mes>')
def get_consolidado_mes(ano, mes):
    db = get_db()
    query = '''
        SELECT s.nome as setor, SUM(p.peso) as total_peso
        FROM pesos p
        JOIN setores s ON p.setor_id = s.id
        JOIN registros_dia r ON p.registro_id = r.id
        WHERE r.data LIKE ?
        GROUP BY s.nome
    '''
    filtro = f'%-{mes}-{ano}'
    rows = db.execute(query, (filtro,)).fetchall()
    return jsonify([dict(row) for row in rows])

@app.route('/api/exportar/<ano>/<mes>')
def exportar_excel(ano, mes):
    db = get_db()
    setores = db.execute("SELECT id, nome FROM setores ORDER BY id").fetchall()
    colunas_horarios = ["05:30", "08:00", "13:00", "16:00", "20:00"]
    dias_mes = monthrange(int(ano), int(mes))[1]
    output = BytesIO()

    # --- Configurações de Estilo ---
    fill_fundo_geral = PatternFill(start_color="00A4A7", end_color="00A4A7", fill_type="solid")
    fill_header_cinza = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    fill_verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fill_azul_claro = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_bold = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def aplicar_layout_base(ws, titulo_superior):
        # Pintar fundo da planilha
        for r in range(1, 100):
            for c in range(1, 20):
                ws.cell(row=r, column=c).fill = fill_fundo_geral
        
        ws.column_dimensions['A'].width = 1.1 # ~13px
        ws.column_dimensions['B'].width = 35
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12

        # Cabeçalhos
        ws.merge_cells("B2:H2")
        ws["B2"].value = titulo_superior
        for cell in ws["2:2"]: # Aplica borda na linha 2
            if cell.column >= 2 and cell.column <= 8:
                cell.fill = fill_header_cinza
                cell.alignment = center_align
                cell.font = font_bold
                cell.border = thin_border

        ws.merge_cells("B3:B4")
        ws["B3"].value = "ITENS"
        ws["B3"].fill = fill_header_cinza
        ws["B3"].alignment = center_align
        ws["B3"].font = font_bold
        ws["B3"].border = thin_border
        ws["B4"].border = thin_border

        ws.merge_cells("C3:G3")
        ws["C3"].value = "HORARIO"
        ws["C3"].fill = fill_verde
        ws["C3"].alignment = center_align
        ws["C3"].font = font_bold
        ws["C3"].border = thin_border

        for i, hora in enumerate(colunas_horarios):
            cell = ws.cell(row=4, column=3+i)
            cell.value = hora
            cell.fill = fill_verde
            cell.alignment = center_align
            cell.font = Font(underline="single", bold=True)
            cell.border = thin_border

        ws.merge_cells("H3:H4")
        ws["H3"].value = "TOTAL ITENS"
        ws["H3"].fill = fill_azul_claro
        ws["H3"].alignment = center_align
        ws["H3"].font = font_bold
        ws["H3"].border = thin_border
        ws["H4"].border = thin_border

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- 1. ABA RESUMO MENSAL ---
        ws_resumo = writer.book.create_sheet("RESUMO MENSAL", 0)
        aplicar_layout_base(ws_resumo, f"TOTAL ACUMULADO - {mes}/{ano}")
        
        query_mensal = """
            SELECT s.nome as setor_nome, h.descricao as hora, SUM(p.peso) as peso_total
            FROM pesos p 
            JOIN setores s ON p.setor_id = s.id 
            JOIN horarios h ON p.horario_id = h.id 
            JOIN registros_dia r ON p.registro_id = r.id 
            WHERE r.data LIKE ?
            GROUP BY s.nome, h.descricao
        """
        dados_mes = db.execute(query_mensal, (f"%-{mes}-{ano}",)).fetchall()
        
        row_idx = 5
        totais_colunas = [0, 0, 0, 0, 0] # Para C, D, E, F, G
        total_geral = 0

        for setor in setores:
            ws_resumo.cell(row=row_idx, column=2, value=setor["nome"]).fill = fill_branco
            ws_resumo.cell(row=row_idx, column=2).border = thin_border
            
            soma_linha = 0
            for i, hora in enumerate(colunas_horarios):
                val = next((d["peso_total"] for d in dados_mes if d["setor_nome"] == setor["nome"] and d["hora"] == hora), 0)
                cell = ws_resumo.cell(row=row_idx, column=3+i, value=val)
                cell.fill = fill_branco
                cell.border = thin_border
                cell.alignment = center_align
                soma_linha += val
                totais_colunas[i] += val
            
            cell_total_l = ws_resumo.cell(row=row_idx, column=8, value=soma_linha)
            cell_total_l.fill = fill_azul_claro
            cell_total_l.font = font_bold
            cell_total_l.border = thin_border
            cell_total_l.alignment = center_align
            total_geral += soma_linha
            row_idx += 1

        # Linha de Totais do Mês
        ws_resumo.cell(row=row_idx, column=2, value="TOTAL GERAL").font = font_bold
        ws_resumo.cell(row=row_idx, column=2).fill = fill_header_cinza
        ws_resumo.cell(row=row_idx, column=2).border = thin_border
        for i, total_col in enumerate(totais_colunas):
            c = ws_resumo.cell(row=row_idx, column=3+i, value=total_col)
            c.fill = fill_header_cinza
            c.font = font_bold
            c.border = thin_border
            c.alignment = center_align
        ws_resumo.cell(row=row_idx, column=8, value=total_geral).fill = fill_azul_claro
        ws_resumo.cell(row=row_idx, column=8).font = font_bold
        ws_resumo.cell(row=row_idx, column=8).border = thin_border

        # --- 2. ABAS DIÁRIAS ---
        for dia in range(1, dias_mes + 1):
            data_str = f"{dia:02d}-{mes}-{ano}"
            ws = writer.book.create_sheet(f"{dia:02d}")
            aplicar_layout_base(ws, f"PESO SUJO - {data_str}")
            
            query_dia = """
                SELECT s.nome as setor_nome, h.descricao as hora, p.peso 
                FROM pesos p 
                JOIN setores s ON p.setor_id = s.id 
                JOIN horarios h ON p.horario_id = h.id 
                JOIN registros_dia r ON p.registro_id = r.id 
                WHERE r.data = ?
            """
            dados_dia = db.execute(query_dia, (data_str,)).fetchall()
            
            row_d = 5
            totais_dia_col = [0, 0, 0, 0, 0]
            total_dia_geral = 0

            for setor in setores:
                ws.cell(row=row_d, column=2, value=setor["nome"]).fill = fill_branco
                ws.cell(row=row_d, column=2).border = thin_border
                
                soma_h = 0
                for i, hora in enumerate(colunas_horarios):
                    val = next((d["peso"] for d in dados_dia if d["setor_nome"] == setor["nome"] and d["hora"] == hora), 0)
                    cell = ws.cell(row=row_d, column=3+i, value=val)
                    cell.fill = fill_branco
                    cell.border = thin_border
                    cell.alignment = center_align
                    soma_h += val
                    totais_dia_col[i] += val
                
                ws.cell(row=row_d, column=8, value=soma_h).fill = fill_azul_claro
                ws.cell(row=row_d, column=8).font = font_bold
                ws.cell(row=row_d, column=8).border = thin_border
                total_dia_geral += soma_h
                row_d += 1

            # Linha de Totais do Dia
            ws.cell(row=row_d, column=2, value="TOTAL DIA").font = font_bold
            ws.cell(row=row_d, column=2).fill = fill_header_cinza
            ws.cell(row=row_d, column=2).border = thin_border
            for i, t_col in enumerate(totais_dia_col):
                c = ws.cell(row=row_d, column=3+i, value=t_col)
                c.fill = fill_header_cinza
                c.font = font_bold
                c.border = thin_border
                c.alignment = center_align
            ws.cell(row=row_d, column=8, value=total_dia_geral).fill = fill_azul_claro
            ws.cell(row=row_d, column=8).font = font_bold
            ws.cell(row=row_d, column=8).border = thin_border

    if 'Sheet' in writer.book.sheetnames:
        writer.book.remove(writer.book['Sheet'])
        
    output.seek(0)
    return send_file(output, download_name=f"Relatorio_{mes}_{ano}.xlsx", as_attachment=True)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
